import os
import zipfile
import xml.etree.ElementTree as ET

def get_xlsx_data(file_path):
    try:
        # Check if openpyxl is available
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        result = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            result.append(f"Sheet: {sheet_name}")
            # print first 10 rows
            for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if r_idx < 15:
                    # Filter out none columns if the whole row is empty
                    if any(val is not None for val in row):
                        result.append(f"  Row {r_idx}: {row[:8]}")
                else:
                    result.append("  ... (truncated)")
                    break
        return "\n".join(result)
    except ImportError:
        return "openpyxl not installed. Attempting XML parsing..."
    except Exception as e:
        return f"Error reading {file_path}: {str(e)}"

if __name__ == "__main__":
    dir_path = r"C:\Desarrollo\mmapp\Temporales\Metricas_App_de_armado"
    files = ["opiniones.xlsx", "opiniones_Abril_20_2024.xlsx", "opiniones_Junio_2025.xlsx"]
    for f in files:
        full_path = os.path.join(dir_path, f)
        print(f"\n====================\nFILE: {f}\n====================")
        print(get_xlsx_data(full_path))
