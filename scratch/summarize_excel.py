import os
import openpyxl

def summarize_xlsx(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    summary = []
    summary.append(f"### Archivo: {os.path.basename(file_path)}")
    for name in wb.sheetnames:
        sheet = wb[name]
        rows = list(sheet.iter_rows(values_only=True))
        non_empty_rows = [r for r in rows if any(val is not None for val in r)]
        summary.append(f"- **Hoja**: `{name}` | Total filas con datos: {len(non_empty_rows)}")
        if not non_empty_rows:
            continue
        headers = non_empty_rows[0]
        summary.append(f"  - **Cabeceras**: {headers}")
        
        # Analyze first sheet feedback if it looks like Hoja1
        if name == "Hoja1" and len(non_empty_rows) > 1:
            q1_counts = {}
            q2_counts = {}
            sentiment = {"Positivas": 0, "Neutrales": 0, "Negativas": 0}
            comments = []
            
            for row in non_empty_rows[1:]:
                # Map headers
                row_dict = dict(zip(headers, row))
                
                # Pregunta 1
                q1 = row_dict.get("Pregunta1")
                if q1:
                    q1_counts[q1] = q1_counts.get(q1, 0) + 1
                    
                # Pregunta 2
                q2 = row_dict.get("Pregunta2")
                if q2:
                    q2_counts[q2] = q2_counts.get(q2, 0) + 1
                
                # Pregunta 3 / Comentarios (if exists)
                q3 = row_dict.get("Pregunta3")
                if q3 and isinstance(q3, str) and len(q3.strip()) > 3:
                    comments.append((row_dict.get("Nombre", "Anónimo"), q3))
                
                # Sentiments in June 2025 sheet
                if "Positivas" in row_dict:
                    if row_dict.get("Positivas") == 1:
                        sentiment["Positivas"] += 1
                    elif row_dict.get("Neutrales") == 1:
                        sentiment["Neutrales"] += 1
                    elif row_dict.get("Negativas") == 1:
                        sentiment["Negativas"] += 1
            
            summary.append("  - **Respuestas Pregunta 1 (¿Qué ayudó más?):**")
            for k, v in q1_counts.items():
                summary.append(f"    - {k}: {v}")
                
            summary.append("  - **Respuestas Pregunta 2 (¿Recomienda el manual/es bueno?):**")
            for k, v in q2_counts.items():
                summary.append(f"    - {k}: {v}")
                
            if any(sentiment.values()):
                summary.append(f"  - **Sentimiento (2025):** Positivas={sentiment['Positivas']}, Neutrales={sentiment['Neutrales']}, Negativas={sentiment['Negativas']}")
            
            if comments:
                summary.append(f"  - **Comentarios destacados (total {len(comments)}):**")
                for name_u, comm in comments[:5]:
                    summary.append(f"    - *{name_u}*: \"{comm}\"")
                if len(comments) > 5:
                    summary.append(f"    - ... y {len(comments)-5} comentarios más.")
                    
        # Analyze Hoja2 for scan data (if present)
        elif name == "Hoja2":
            summary.append("  - **Datos de Escaneos (Hoja 2):**")
            for row in non_empty_rows:
                # print any row that contains numbers or totals
                row_str = " | ".join(str(x) for x in row if x is not None)
                if any(kw in row_str.lower() for kw in ["escritorio", "computo", "armario", "escaneo", "total"]):
                    summary.append(f"    - {row_str}")
                    
    return "\n".join(summary)

if __name__ == "__main__":
    dir_path = r"C:\Desarrollo\mmapp\Temporales\Metricas_App_de_armado"
    files = ["opiniones.xlsx", "opiniones_Abril_20_2024.xlsx", "opiniones_Junio_2025.xlsx"]
    for f in files:
        print(summarize_xlsx(os.path.join(dir_path, f)))
        print("\n" + "="*40 + "\n")
